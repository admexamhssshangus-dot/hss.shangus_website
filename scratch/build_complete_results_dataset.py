import openpyxl, sys, json, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
sys.stdout.reconfigure(encoding='utf-8')

print("Starting Consolidated JKBOSE & TC/DC Result Enrichment Dataset Generation...")

# 1. Load provisonal cum character.xlsm
wb_prov = openpyxl.load_workbook(r'D:\Shk_Gulfam\Projects\context\provisonal cum character.xlsm', data_only=True)
s_dig2 = wb_prov['db_digital2']
rows_dig2 = list(s_dig2.iter_rows(values_only=True))

s_rev = wb_prov['reverse_engineering']
rows_rev = list(s_rev.iter_rows(values_only=True))
wb_prov.close()

prov_data_by_reg = {}
prov_data_by_exam_roll = {}
prov_data_by_adm_no = {}
prov_data_by_name_father = {}

# Process db_digital2 (Certf. No. 1001 to 1367)
for r in rows_dig2[1:]:
    reg = str(r[11] or '').strip()
    exam_roll = str(r[24] or '').strip()
    adm_no = str(r[10] or '').strip()
    name = str(r[12] or '').strip()
    father = str(r[15] or '').strip()
    
    cert_no = str(r[7] or '').strip()
    wd_date = str(r[41] or '').strip()
    issue_date = str(r[43] or '').strip()
    if isinstance(r[41], datetime):
        wd_date = r[41].strftime('%d-%m-%Y')
    if isinstance(r[43], datetime):
        issue_date = r[43].strftime('%d-%m-%Y')
    
    cc_dc = str(r[44] or '').strip()
    if cc_dc and not ('dated' in cc_dc.lower()):
        d = issue_date or wd_date
        if d:
            cc_dc = f"{cc_dc} dated {d}"

    entry = {
        'source': 'db_digital2',
        'cert_no': cert_no,
        'adm_no': adm_no,
        'adm_date': str(r[9] or '').strip(),
        'reg_no': reg,
        'name': name,
        'father': father,
        'mother': str(r[16] or '').strip(),
        'village': str(r[17] or '').strip(),
        'tehsil': str(r[18] or '').strip(),
        'district': str(r[19] or '').strip(),
        'cls': str(r[20] or '').strip(),
        'session': str(r[23] or '').strip(),
        'exam_roll': exam_roll,
        'result': str(r[25] or '').strip() or 'Passed',
        'division': str(r[26] or '').strip(),
        'marks': str(r[28] or '').strip(),
        'dob': str(r[36] or '').strip(),
        'dob_words': str(r[37] or '').strip(),
        'wd_date': wd_date,
        'issue_date': issue_date,
        'cc_dc_no': cc_dc
    }
    
    if reg:
        prov_data_by_reg[reg] = entry
    if exam_roll:
        prov_data_by_exam_roll[exam_roll] = entry
    if adm_no:
        prov_data_by_adm_no[adm_no] = entry
    if name:
        key = f"{name.lower()}_{father.lower()}"
        prov_data_by_name_father[key] = entry

# Process reverse_engineering (Certf. No. 853 to 970)
for r in rows_rev[1:]:
    reg = str(r[2] or '').strip()
    exam_roll = str(r[16] or '').strip()
    adm_no = str(r[4] or '').strip()
    name = str(r[5] or '').strip()
    father = str(r[11] or '').strip()
    
    cert_no = str(r[24] or '').strip()
    wd_date = str(r[22] or '').strip()
    issue_date = str(r[23] or '').strip()
    if isinstance(r[22], datetime):
        wd_date = r[22].strftime('%d-%m-%Y')
    if isinstance(r[23], datetime):
        issue_date = r[23].strftime('%d-%m-%Y')
        
    cc_dc = cert_no
    if cc_dc and not ('dated' in cc_dc.lower()):
        d = issue_date or wd_date
        if d:
            cc_dc = f"{cc_dc} dated {d}"

    entry = {
        'source': 'reverse_engineering',
        'cert_no': cert_no,
        'adm_no': adm_no,
        'adm_date': str(r[3] or '').strip(),
        'reg_no': reg,
        'name': name,
        'father': father,
        'mother': str(r[12] or '').strip(),
        'village': str(r[13] or '').strip(),
        'cls': str(r[14] or '').strip(),
        'session': str(r[15] or '').strip(),
        'exam_roll': exam_roll,
        'result': str(r[17] or '').strip() or 'Passed',
        'division': str(r[17] or '').strip(),
        'marks': str(r[18] or '').strip(),
        'dob': str(r[20] or '').strip(),
        'dob_words': str(r[21] or '').strip(),
        'wd_date': wd_date,
        'issue_date': issue_date,
        'cc_dc_no': cc_dc
    }

    if reg and reg not in prov_data_by_reg:
        prov_data_by_reg[reg] = entry
    if exam_roll and exam_roll not in prov_data_by_exam_roll:
        prov_data_by_exam_roll[exam_roll] = entry
    if adm_no and adm_no not in prov_data_by_adm_no:
        prov_data_by_adm_no[adm_no] = entry

print(f"Indexed {len(prov_data_by_reg)} records from provisonal cum character.xlsm")

# 2. Load db_30 Jul 2026.xlsx (source_data)
wb_src = openpyxl.load_workbook(r'D:\Shk_Gulfam\Projects\hss_shangus\db_30 Jul 2026.xlsx', read_only=True, data_only=True)
s_src = wb_src['source_data']
rows_src = list(s_src.iter_rows(values_only=True))
wb_src.close()

consolidated_records = []
matched_count = 0

for idx, r in enumerate(rows_src[1:]):
    form_no = str(r[1] or '').strip()
    class_roll = str(r[2] or '').strip()
    adm_date = str(r[4] or '').strip()
    adm_no = str(r[5] or '').strip()
    cls = str(r[6] or '').strip()
    session = str(r[7] or '').strip()
    reg_no = str(r[8] or '').strip()
    name = str(r[10] or '').strip()
    father = str(r[11] or '').strip()
    mother = str(r[12] or '').strip()
    dob_fig = str(r[13] or '').strip()
    dob_wrd = str(r[14] or '').strip()
    village = str(r[15] or '').strip()
    stream = str(r[27] or '').strip()

    exam_mode = str(r[60] or '').strip()
    exam_roll = str(r[61] or '').strip()
    result = str(r[62] or '').strip()
    marks = str(r[63] or '').strip()
    wd_date = str(r[64] or '').strip()
    cc_dc = str(r[65] or '').strip()
    remarks = str(r[66] or '').strip()

    # Match against provisonal cum character
    name_key = f"{name.lower()}_{father.lower()}"
    prov = (
        prov_data_by_reg.get(reg_no) or
        prov_data_by_exam_roll.get(exam_roll) or
        prov_data_by_adm_no.get(adm_no) or
        prov_data_by_name_father.get(name_key)
    )

    if prov:
        matched_count += 1
        if not exam_roll and prov.get('exam_roll'):
            exam_roll = prov['exam_roll']
        if not result and prov.get('result'):
            result = prov['result']
        if not marks and prov.get('marks'):
            marks = prov['marks']
        if not wd_date and prov.get('wd_date'):
            wd_date = prov['wd_date']
        if not cc_dc and prov.get('cc_dc_no'):
            cc_dc = prov['cc_dc_no']
        if not exam_mode and prov.get('session'):
            exam_mode = prov['session']

    # Division calculation if marks present
    div_distinc = ''
    if marks:
        num_m = None
        import re
        m_match = re.search(r'(\d+)(?:\s*/\s*(\d+))?', str(marks))
        if m_match:
            obt = int(m_match.group(1))
            tot = int(m_match.group(2)) if m_match.group(2) else 500
            pct = (obt / tot) * 100
            if pct >= 75: div_distinc = 'Distinction'
            elif pct >= 60: div_distinc = '1st Division'
            elif pct >= 45: div_distinc = '2nd Division'
            else: div_distinc = '3rd Division'

    consolidated_records.append({
        'sNo': idx + 1,
        'formNo': form_no,
        'classRollNo': class_roll,
        'regNo': reg_no,
        'studentName': name,
        'fatherName': father,
        'motherName': mother,
        'className': cls,
        'stream': stream,
        'session': session,
        'examMode': exam_mode or 'Annual Regular 2025 (Oct.-Nov.)',
        'examRollNo': exam_roll,
        'result': result or ('Passed' if marks or exam_roll else ''),
        'marksReapp': marks,
        'divDistinc': div_distinc,
        'withdrawalDate': wd_date,
        'ccDcNo': cc_dc,
        'remarks': remarks
    })

print(f"Total source records processed: {len(consolidated_records)}")
print(f"Total matched & enriched with provisonal cum character: {matched_count}")

# 3. Create enriched XLSX workbook
out_wb = openpyxl.Workbook()
ws = out_wb.active
ws.title = "JKBOSE_Master_Results"

headers = [
    'S.No.',
    'Form No.',
    'Class R.No.',
    'Board Reg. No.',
    "Student's Name",
    "Father's Name",
    'Class',
    'Stream',
    'Session',
    'Exam Mode (Current)',
    'Exam R.No. (Current)',
    'Result (Current)',
    'Marks/Reapp (Current)',
    'Div/Distinc (Current)',
    'Date of withdrawl/result',
    'No. & Date of CC/DC Issued (This Institution)',
    'Remarks'
]

ws.append(headers)

# Styling header
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # Navy
thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Append data rows with explicit Text format (@) to prevent scientific notation
for row_idx, rec in enumerate(consolidated_records, start=2):
    row_vals = [
        str(rec['sNo']),
        str(rec['formNo']),
        str(rec['classRollNo']),
        str(rec['regNo']),
        str(rec['studentName']),
        str(rec['fatherName']),
        str(rec['className']),
        str(rec['stream']),
        str(rec['session']),
        str(rec['examMode']),
        str(rec['examRollNo']),
        str(rec['result']),
        str(rec['marksReapp']),
        str(rec['divDistinc']),
        str(rec['withdrawalDate']),
        str(rec['ccDcNo']),
        str(rec['remarks'])
    ]
    for col_idx, val in enumerate(row_vals, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.number_format = '@' # Explicit Text format

# Column widths
col_widths = [6, 12, 12, 22, 24, 24, 8, 14, 12, 28, 18, 16, 22, 18, 22, 34, 20]
for idx, w in enumerate(col_widths, 1):
    col_letter = openpyxl.utils.get_column_letter(idx)
    ws.column_dimensions[col_letter].width = w

out_path = r'D:\Shk_Gulfam\Projects\hss_shangus\scratch\JKBOSE_Complete_Results_Enriched.xlsx'
out_wb.save(out_path)
out_wb.close()

print(f"Successfully generated complete enriched file at: {out_path}")
