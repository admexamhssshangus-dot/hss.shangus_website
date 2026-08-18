import openpyxl, sys, json, os
sys.stdout.reconfigure(encoding='utf-8')

# Read source_data
wb1 = openpyxl.load_workbook(r'D:\Shk_Gulfam\Projects\hss_shangus\db_30 Jul 2026.xlsx', read_only=True, data_only=True)
s1 = wb1['source_data']
rows1 = list(s1.iter_rows(values_only=True))
wb1.close()

# Read provisonal cum character.xlsm
wb2 = openpyxl.load_workbook(r'D:\Shk_Gulfam\Projects\context\provisonal cum character.xlsm', data_only=True)
s2 = wb2['db_digital2']
rows2 = list(s2.iter_rows(values_only=True))
s3 = wb2['reverse_engineering']
rows3 = list(s3.iter_rows(values_only=True))
wb2.close()

# Collect provisonal data by Reg No and by Exam Roll No
prov_by_reg = {}
prov_by_exam_roll = {}

for r in rows2[1:]:
    reg = str(r[11] or '').strip()
    exam_roll = str(r[24] or '').strip()
    if reg:
        entry = {
            'source': 'db_digital2',
            'reg': reg,
            'cert_no': str(r[7] or '').strip(),
            'adm_date': str(r[9] or '').strip(),
            'adm_no': str(r[10] or '').strip(),
            'name': str(r[12] or '').strip(),
            'father': str(r[15] or '').strip(),
            'mother': str(r[16] or '').strip(),
            'cls': str(r[20] or '').strip(),
            'session': str(r[23] or '').strip(),
            'exam_roll': exam_roll,
            'result': str(r[25] or '').strip(),
            'division': str(r[26] or '').strip(),
            'marks': str(r[28] or '').strip(),
            'dob': str(r[36] or '').strip(),
            'dob_words': str(r[37] or '').strip(),
            'wd_date': str(r[41] or '').strip(),
            'issue_date': str(r[43] or '').strip(),
            'cc_dc_no': str(r[44] or '').strip()
        }
        prov_by_reg[reg] = entry
        if exam_roll:
            prov_by_exam_roll[exam_roll] = entry

for r in rows3[1:]:
    reg = str(r[2] or '').strip()
    exam_roll = str(r[16] or '').strip()
    if reg and reg not in prov_by_reg:
        entry = {
            'source': 'reverse_engineering',
            'reg': reg,
            'adm_date': str(r[3] or '').strip(),
            'adm_no': str(r[4] or '').strip(),
            'name': str(r[5] or '').strip(),
            'father': str(r[11] or '').strip(),
            'mother': str(r[12] or '').strip(),
            'cls': str(r[14] or '').strip(),
            'session': str(r[15] or '').strip(),
            'exam_roll': exam_roll,
            'result': str(r[17] or '').strip(),
            'passed': str(r[19] or '').strip(),
            'marks': str(r[18] or '').strip(),
            'dob': str(r[20] or '').strip(),
            'dob_words': str(r[21] or '').strip(),
            'wd_date': str(r[22] or '').strip(),
            'issue_date': str(r[23] or '').strip(),
            'cc_dc_no': str(r[24] or '').strip()
        }
        prov_by_reg[reg] = entry
        if exam_roll:
            prov_by_exam_roll[exam_roll] = entry
    elif reg in prov_by_reg:
        existing = prov_by_reg[reg]
        if not existing.get('wd_date') and r[22]:
            existing['wd_date'] = str(r[22]).strip()
        if not existing.get('cc_dc_no') and r[24]:
            existing['cc_dc_no'] = str(r[24]).strip()

print(f'Total unique students indexed in provisonal cum character.xlsm: {len(prov_by_reg)}')

# Now match and merge with source_data
matched_count = 0
updated_from_prov = 0
merged_records = []

for idx, r in enumerate(rows1[1:]):
    form_no = str(r[1] or '').strip()
    reg_no = str(r[8] or '').strip()
    name = str(r[10] or '').strip()
    father = str(r[11] or '').strip()
    cls = str(r[6] or '').strip()
    session = str(r[7] or '').strip()
    
    exam_mode = str(r[60] or '').strip()
    exam_roll = str(r[61] or '').strip()
    result = str(r[62] or '').strip()
    marks = str(r[63] or '').strip()
    wd_date = str(r[64] or '').strip()
    cc_dc = str(r[65] or '').strip()
    remarks = str(r[66] or '').strip()
    
    prov_match = prov_by_reg.get(reg_no) or (prov_by_exam_roll.get(exam_roll) if exam_roll else None)
    
    if prov_match:
        matched_count += 1
        if not exam_roll and prov_match.get('exam_roll'):
            exam_roll = str(prov_match['exam_roll'])
        if not result and prov_match.get('result'):
            result = str(prov_match['result'])
            if prov_match.get('division'):
                result += f' ({prov_match["division"]})'
        if not marks and prov_match.get('marks'):
            marks = str(prov_match['marks'])
        if not wd_date and prov_match.get('wd_date'):
            wd_date = str(prov_match['wd_date'])
        if not cc_dc and prov_match.get('cc_dc_no'):
            cc_dc = f"{prov_match['cc_dc_no']}".strip()
            if prov_match.get('issue_date') or prov_match.get('wd_date'):
                cc_dc += f" dated {prov_match.get('issue_date') or prov_match.get('wd_date')}"
        updated_from_prov += 1
        
    if exam_roll or result or marks or wd_date or cc_dc:
        merged_records.append({
            'rowIdx': idx + 2,
            'formNo': form_no,
            'regNo': reg_no,
            'name': name,
            'father': father,
            'cls': cls,
            'session': session,
            'examMode': exam_mode,
            'examRollNo': exam_roll,
            'result': result,
            'marks': marks,
            'withdrawalDate': wd_date,
            'ccDcNo': cc_dc,
            'remarks': remarks
        })

print(f'Total merged records with exam/result/TC info: {len(merged_records)}')
print(f'Matched with provisonal cum character: {matched_count}')

with open('scratch/merged_records_sample.json', 'w', encoding='utf-8') as f:
    json.dump(merged_records[:20], f, indent=2)

print('Sample 5 records:')
for m in merged_records[:5]:
    print(m)
