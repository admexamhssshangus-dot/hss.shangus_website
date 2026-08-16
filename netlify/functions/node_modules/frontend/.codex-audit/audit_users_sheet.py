import json
import re
import sys
from collections import Counter, defaultdict

from openpyxl import load_workbook


def norm(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def password_format(value):
    if value is None or str(value) == "":
        return "missing"
    text = str(value)
    if re.fullmatch(r"\$2[aby]\$\d{2}\$[./A-Za-z0-9]{53}", text):
        return "bcrypt_hash"
    if re.match(r"^\$argon2(?:id|i|d)\$", text):
        return "argon2_hash"
    if re.fullmatch(r"[a-fA-F0-9]{32}", text):
        return "md5_like_hash"
    if re.fullmatch(r"[a-fA-F0-9]{40}", text):
        return "sha1_like_hash"
    if re.fullmatch(r"[a-fA-F0-9]{64}", text):
        return "sha256_like_hash"
    if re.fullmatch(r"[a-fA-F0-9]{128}", text):
        return "sha512_like_hash"
    return "plaintext_or_temporary"


path = sys.argv[1]
workbook = load_workbook(path, read_only=True, data_only=True)
sheet = workbook["Users"]

candidate_rows = []
for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 25), values_only=True), 1):
    values = list(row)
    normalized = [norm(value) for value in values]
    score = sum(
        1
        for item in normalized
        if item in {
            "email", "emailaddress", "useremail", "username", "password", "passwordplain", "pass",
            "userpassword", "loginpassword", "role", "userrole", "accounttype",
            "name", "fullname", "displayname", "studentname", "teachername",
        }
    )
    candidate_rows.append((score, row_number, values, normalized))

score, header_row, raw_headers, headers = max(candidate_rows, key=lambda item: (item[0], -item[1]))
if score == 0:
    raise RuntimeError("Could not identify a Users-sheet header row in the first 25 rows")


def find_col(*names):
    wanted = {norm(name) for name in names}
    return next((index for index, header in enumerate(headers) if header in wanted), -1)


email_col = find_col("email", "email address", "user email", "username")
password_col = find_col("password", "password plain", "PasswordPlain", "pass", "user password", "login password")
role_col = find_col("role", "user role", "account type")
name_col = find_col("name", "full name", "display name", "student name", "teacher name")
email_re = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")

rows = []
email_rows = defaultdict(list)
formats = Counter()
password_quality = Counter()
roles = Counter()
email_domains = Counter()

for excel_row, values in enumerate(
    sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row, values_only=True),
    header_row + 1,
):
    values = list(values)
    if not any(value is not None and str(value).strip() for value in values):
        continue
    email = str(values[email_col] or "").strip().lower() if email_col >= 0 else ""
    fmt = password_format(values[password_col]) if password_col >= 0 else "column_missing"
    role = str(values[role_col] or "").strip() if role_col >= 0 else ""
    name = str(values[name_col] or "").strip() if name_col >= 0 else ""
    formats[fmt] += 1
    password_text = str(values[password_col] or "") if password_col >= 0 else ""
    if fmt == "plaintext_or_temporary":
        if len(password_text) < 6:
            password_quality["firebase_rejected_under_6"] += 1
        elif len(password_text) < 12:
            password_quality["accepted_by_firebase_but_under_site_policy_12"] += 1
        else:
            password_quality["length_12_or_more"] += 1
        if password_text.isdigit():
            password_quality["numeric_only"] += 1
        if password_text.lower() == email.split("@", 1)[0].lower():
            password_quality["same_as_email_local_part"] += 1
    roles[role or "(missing)"] += 1
    if "@" in email:
        email_domains[email.rsplit("@", 1)[1]] += 1
    item = {
        "excelRow": excel_row,
        "email": email,
        "emailValid": bool(email_re.fullmatch(email)),
        "passwordFormat": fmt,
        "passwordLength": len(str(values[password_col] or "")) if password_col >= 0 else 0,
        "role": role or None,
        "hasDisplayName": bool(name),
    }
    rows.append(item)
    if email:
        email_rows[email].append(excel_row)

result = {
    "sheetNames": workbook.sheetnames,
    "users": {
        "headerRow": header_row,
        "usedRowCount": sheet.max_row,
        "usedColumnCount": sheet.max_column,
        "headers": [str(value or "").strip() for value in raw_headers],
        "detectedColumns": {
            "emailIndex": email_col,
            "passwordIndex": password_col,
            "roleIndex": role_col,
            "nameIndex": name_col,
        },
        "dataRowCount": len(rows),
        "passwordFormats": dict(formats),
        "passwordQuality": dict(password_quality),
        "roles": dict(roles),
        "emailDomains": dict(email_domains),
        "suspiciousEmailDomainRows": [
            {"excelRow": row["excelRow"], "email": row["email"]}
            for row in rows
            if "@" in row["email"] and row["email"].rsplit("@", 1)[1]
            in {"gamil.com", "gmial.com", "gmail.co", "gmail.con", "gmail.cim", "gmail.om"}
        ],
        "duplicates": [
            {"email": email, "excelRows": excel_rows}
            for email, excel_rows in email_rows.items()
            if len(excel_rows) > 1
        ],
        "invalidEmailRows": [
            {"excelRow": row["excelRow"], "email": row["email"]}
            for row in rows
            if not row["emailValid"]
        ],
        "missingPasswordRows": [
            {"excelRow": row["excelRow"], "email": row["email"]}
            for row in rows
            if row["passwordFormat"] == "missing"
        ],
        "firebaseRejectedPasswordRows": [
            {"excelRow": row["excelRow"], "email": row["email"], "passwordLength": row["passwordLength"]}
            for row in rows
            if row["passwordFormat"] == "plaintext_or_temporary" and row["passwordLength"] < 6
        ],
        "duplicateCredentialConflicts": [
            {
                "email": email,
                "excelRows": excel_rows,
                "passwordsDiffer": len({
                    str(sheet.cell(row=excel_row, column=password_col + 1).value or "")
                    for excel_row in excel_rows
                }) > 1,
            }
            for email, excel_rows in email_rows.items()
            if len(excel_rows) > 1
        ],
        "conflictingDuplicateRoles": [
            {
                "email": email,
                "excelRows": excel_rows,
                "roles": sorted({row["role"] for row in rows if row["email"] == email}),
            }
            for email, excel_rows in email_rows.items()
            if len(excel_rows) > 1 and len({row["role"] for row in rows if row["email"] == email}) > 1
        ],
    },
}

print(json.dumps(result, indent=2))
