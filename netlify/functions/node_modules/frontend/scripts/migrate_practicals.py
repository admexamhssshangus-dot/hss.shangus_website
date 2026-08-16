"""
migrate_practicals.py
Reads 'practical_data' sheet from 'db_30 Jul 2026.xlsx'.
Dynamically detects and parses headers per section (Row 1 for 11th, Row 16 for 12th, Row 31 for 11th ext).
Cross-references 'source_data' sheet to enrich each practical record with:
  - boardRegNo  (Board Reg. No.)
  - stream      (Stream)
  - subjects    (Subs)
  Matching key: classRollNo + class + session (2024-25)
"""

import openpyxl, re, json

EXCEL_PATH = 'db_30 Jul 2026.xlsx'
OUTPUT_JS  = 'src/data/cleanPracticalsSeedData.js'

SUBJECT_CODES = {
    'botany': 'BO', 'zoology': 'ZO', 'biology': 'BI',
    'physics': 'PH', 'chemistry': 'CH', 'mathematics': 'MA',
    'math': 'MA', 'urdu': 'UR', 'education': 'ED',
    'history': 'HT', 'political': 'PS', 'economics': 'EC',
    'environmental': 'ES', 'physical education': 'PD',
    'healthcare': 'HTC', 'it and ites': 'ITE', 'ite': 'ITE',
    'general english': 'EN', 'english': 'EN',
}

SUBJECT_NAMES = {
    'BO': 'Botany', 'ZO': 'Zoology', 'BI': 'Biology (Botany & Zoology)',
    'PH': 'Physics', 'CH': 'Chemistry', 'MA': 'Mathematics',
    'UR': 'Urdu', 'ED': 'Education', 'HT': 'History',
    'PS': 'Political Science', 'EC': 'Economics', 'ES': 'Environmental Science',
    'PD': 'Physical Education', 'HTC': 'Healthcare', 'ITE': 'IT and ITES',
    'EN': 'General English',
}

def get_subject_code(subject_str):
    if not subject_str:
        return 'XX'
    s = subject_str.lower().strip()
    m = re.search(r'\(([A-Z]{2,4})\)', subject_str)
    if m:
        code = m.group(1).upper()
        if code in SUBJECT_NAMES:
            return code
    for kw, code in SUBJECT_CODES.items():
        if kw in s:
            return code
    return 'XX'

def normalize_session(session_str):
    if not session_str:
        return '2024-25 (Oct-Nov)'
    s = str(session_str).strip()
    if re.match(r'\d{4}-\d{2}\s*\(', s):
        return s
    if 'regular' in s.lower() and '2025' in s and 'oct' in s.lower():
        return '2024-25 (Oct-Nov)'
    if 'oct' in s.lower() or 'nov' in s.lower():
        return '2024-25 (Oct-Nov)'
    return s

def normalize_session_key(session_str):
    s = str(session_str or '').strip()
    m = re.match(r'(\d{4}-\d{2})', s)
    if m:
        return m.group(1)
    return s

def make_doc_id(cls, subj_code, prac_type, session_text):
    session_slug = session_text.lower()
    session_slug = re.sub(r'[^a-z0-9]+', '-', session_slug).strip('-')
    pt = (prac_type or 'internal').lower()
    return '%sth_%s_%s_%s' % (cls.lower().replace('th', ''), subj_code.lower(), pt, session_slug)

def parse_student_col(col_header):
    if not col_header or str(col_header).strip() in ('None', '', 'none'):
        return None
    col_header = str(col_header).strip()
    m = re.match(r'^(\d+)\s*/\s*(\d+)\.\s*(.+?)\s*\((.+?)\)\s*$', col_header)
    if m:
        s_no = int(m.group(1))
        exam_roll = m.group(2).strip()
        name = m.group(3).strip()
        parent = m.group(4).strip()
        return (s_no, str(s_no), exam_roll, name, parent)
    m2 = re.match(r'^(\d+)\s*/\s*(\d+)\.\s*(.+?)\s*$', col_header)
    if m2:
        s_no = int(m2.group(1))
        exam_roll = m2.group(2).strip()
        name = m2.group(3).strip()
        return (s_no, str(s_no), exam_roll, name, '')
    return None

def build_source_lookup(ws_source):
    rows = ws_source.iter_rows(min_row=1, values_only=True)
    headers = [str(h).strip() if h else '' for h in next(rows)]

    def col(row, name_options):
        for name in name_options:
            try:
                idx = headers.index(name)
                v = row[idx]
                return str(v).strip() if v is not None else ''
            except ValueError:
                pass
        return ''

    lookup_by_classroll = {}  # (classRollNo, classNorm, sessionYear) -> info
    lookup_by_examroll  = {}  # (examRollNo, classNorm, sessionYear) -> info

    for row in rows:
        if not any(row):
            continue

        class_roll = col(row, ['Class R.No.', 'classRollNo', 'Class Roll No', 'Class Roll No.'])
        class_val  = col(row, ['Class', 'class'])
        session    = col(row, ['Session', 'session'])
        board_reg  = col(row, ['Board Reg. No.', 'Board Registration Number', 'Board Registration No.'])
        name       = col(row, ["Student's Name", 'studentName', 'name'])
        father     = col(row, ["Father's Name", 'fatherName'])
        stream     = col(row, ['Stream', 'stream'])
        subs       = col(row, ['Subs', 'subs'])
        exam_curr  = col(row, ['Exam R.No. (Current)', 'Exam Roll No. (Current)', 'Exam R.No.', 'Exam Roll No', 'Exam Roll No.'])
        if not subs:
            subs_cols = ['Subjects1', 'Subjects2', 'Subjects3', 'Subjects4', 'Subjects5', 'Subject6']
            subjects_list = [col(row, [sc]) for sc in subs_cols if col(row, [sc])]
            subs = ', '.join(s for s in subjects_list if s and s.lower() != 'none')

        if not class_roll or class_roll.lower() in ('none', '', 'n/a', '-'):
            continue

        class_norm = class_val.lower().strip()
        if class_norm.isdigit():
            class_norm = class_norm + 'th'

        session_year = normalize_session_key(session)

        info = {
            'boardRegNo': board_reg,
            'classRollNo': class_roll,
            'name': name,
            'parentName': father,
            'stream': stream,
            'subjects': subs,
            'class': class_norm,
            'session': session_year,
            'examRollNo': exam_curr
        }

        key_roll = (str(class_roll).strip(), class_norm, session_year)
        lookup_by_classroll[key_roll] = info

        if exam_curr:
            key_exam = (str(exam_curr).strip(), class_norm, session_year)
            lookup_by_examroll[key_exam] = info

    print('Built source_data lookup: %d class-roll entries, %d exam-roll entries' % (len(lookup_by_classroll), len(lookup_by_examroll)))
    return lookup_by_classroll, lookup_by_examroll

def main():
    print('Opening %s...' % EXCEL_PATH)
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    print('Reading source_data for cross-referencing...')
    lookup_by_classroll, lookup_by_examroll = build_source_lookup(wb['source_data'])

    ws = wb['practical_data']

    docs_map = {}
    current_student_defs = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row or not any(row):
            continue

        first_col = str(row[0]).strip().lower() if row[0] is not None else ''

        # Dynamic Header Row Detection
        if first_col == 'timestamp' or 'timestamp' in first_col or (len(row) > 1 and 'email' in str(row[1]).lower()):
            current_student_defs = [parse_student_col(c) if i >= 7 else None for i, c in enumerate(row)]
            valid_students = [s for s in current_student_defs if s]
            print(f'Header Section at Row {row_idx}: {len(valid_students)} students defined. (First: {valid_students[0][3] if valid_students else "N/A"})')
            continue

        class_name  = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
        subject_str = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''

        if not class_name or not subject_str or not any(d.isdigit() for d in class_name):
            continue

        timestamp     = str(row[0]).strip() if row[0] is not None else ''
        teacher_email = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
        teacher_name  = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
        prac_type     = str(row[5]).strip() if len(row) > 5 and row[5] is not None else 'internal'
        session_raw   = str(row[6]).strip() if len(row) > 6 and row[6] is not None else '2024-25 (Oct-Nov)'

        subj_code    = get_subject_code(subject_str)
        subj_name    = SUBJECT_NAMES.get(subj_code, subject_str)
        session_text = normalize_session(session_raw)
        session_year = normalize_session_key(session_text)
        class_norm   = class_name.lower().strip()
        if class_norm.isdigit():
            class_norm += 'th'

        doc_id = make_doc_id(class_name, subj_code, prac_type, session_text)

        if doc_id not in docs_map:
            docs_map[doc_id] = {
                'id': doc_id,
                'className': class_name,
                'subjectCode': subj_code,
                'subjectName': subj_name,
                'teacherEmail': teacher_email,
                'teacherName': teacher_name,
                'practicalType': prac_type,
                'sessionText': session_text,
                'timestamp': timestamp,
                'records': []
            }

        existing_rolls = {r['examRollNo'] for r in docs_map[doc_id]['records']}

        for col_i, sdef in enumerate(current_student_defs):
            if sdef is None or col_i >= len(row):
                continue
            mark_val = row[col_i]
            if mark_val is None:
                continue
            mark_str = str(mark_val).strip()
            if not mark_str or mark_str.lower() in ('none', 'null', 'undefined'):
                continue

            s_no, class_roll, exam_roll, name, parent = sdef

            if exam_roll in existing_rolls:
                for ex in docs_map[doc_id]['records']:
                    if ex['examRollNo'] == exam_roll and not ex.get('totalMarks'):
                        ex['totalMarks'] = mark_str
                        ex['practicalMarks'] = mark_str
                        break
                continue

            # Lookup source_data info by (classRollNo, class, session) or (examRollNo, class, session)
            src_key_roll = (str(s_no), class_norm, session_year)
            src_key_exam = (str(exam_roll), class_norm, session_year)
            src_info = lookup_by_classroll.get(src_key_roll) or lookup_by_examroll.get(src_key_exam) or {}

            board_reg  = src_info.get('boardRegNo', '')
            stream     = src_info.get('stream', '')
            subjects   = src_info.get('subjects', '')
            src_name   = src_info.get('name', '') or name
            src_parent = src_info.get('parentName', '') or parent

            record = {
                'sNo': s_no,
                'classRollNo': class_roll,
                'examRollNo': exam_roll,
                'name': src_name,
                'parentName': src_parent,
                'practicalMarks': mark_str,
                'totalMarks': mark_str,
            }
            if board_reg:
                record['boardRegNo'] = board_reg
            if stream:
                record['stream'] = stream
            if subjects:
                record['subjects'] = subjects

            docs_map[doc_id]['records'].append(record)
            existing_rolls.add(exam_roll)

        if teacher_email:
            docs_map[doc_id]['teacherEmail'] = teacher_email
        if teacher_name:
            docs_map[doc_id]['teacherName'] = teacher_name

    docs = list(docs_map.values())
    for d in docs:
        d['records'].sort(key=lambda r: r.get('sNo', 0))

    enriched_count = sum(1 for d in docs for r in d['records'] if r.get('boardRegNo'))
    total_records  = sum(len(d['records']) for d in docs)

    print('\nParsed %d subject submission documents:' % len(docs))
    for d in docs:
        enriched = sum(1 for r in d['records'] if r.get('boardRegNo'))
        print('  %s (%s): %d records (%d enriched with boardRegNo)' % (
            d['id'], d['className'], len(d['records']), enriched))

    print('\nTotal docs: %d, Total records: %d, Enriched: %d/%d' % (
        len(docs), total_records, enriched_count, total_records))

    js_content = 'export const CLEAN_PRACTICALS_SEED_DATA = '
    js_content += json.dumps(docs, indent=2, ensure_ascii=False)
    js_content += ';\n'

    with open(OUTPUT_JS, 'w', encoding='utf-8') as f:
        f.write(js_content)

    print('Written to %s' % OUTPUT_JS)

if __name__ == '__main__':
    main()
