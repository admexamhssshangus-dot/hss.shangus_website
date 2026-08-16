import json
import re
import sys
from openpyxl import load_workbook

path = sys.argv[1]
wb_values = load_workbook(path, read_only=True, data_only=True)
users = wb_values["Users"]

targets = {
    "shahnawaz": [4, 11],
    "mwani": [19, 308],
    "email_typo": [538],
}

def sanitized_row(row_number):
    values = [users.cell(row_number, col).value for col in range(1, 9)]
    return {
        "row": row_number,
        "email": values[0],
        "name": values[1],
        "passwordPresent": values[2] not in (None, ""),
        "passwordLength": len(str(values[2] or "")),
        "role": values[3],
        "mobilePresent": values[4] not in (None, ""),
        "updatedOn": str(values[5]) if values[5] is not None else None,
        "class": values[6],
        "subject": values[7],
    }

rows = {name: [sanitized_row(r) for r in row_numbers] for name, row_numbers in targets.items()}

wb_formulas = load_workbook(path, read_only=True, data_only=False)
formula_refs = []
pattern = re.compile(r"(?:'Users'|Users)!\$?[A-H]\$?(?:4|19)(?!\d)", re.IGNORECASE)
for sheet in wb_formulas.worksheets:
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("=") and pattern.search(cell.value):
                formula_refs.append({"sheet": sheet.title, "cell": cell.coordinate, "formula": cell.value})

print(json.dumps({"rows": rows, "formulaReferencesToRowsBeingCleared": formula_refs}, indent=2, default=str))
